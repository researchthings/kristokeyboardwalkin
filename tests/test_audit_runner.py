"""Tests for audit/runner.py and audit/report.py end-to-end."""

from __future__ import annotations

import json
from contextlib import closing
from pathlib import Path

import duckdb
from openpyxl import load_workbook

from keywalk_audit.audit import audit, write_json, write_xlsx
from keywalk_audit.fuzzy.geometric_minhash import GeometricMinHashIndex
from keywalk_audit.fuzzy.string_minhash import StringMinHashIndex
from keywalk_audit.hashing.computer import compute_hash
from keywalk_audit.layouts import QWERTY_US
from keywalk_audit.rainbow.builder import geom_lsh_path, string_lsh_path
from keywalk_audit.rainbow.schema import init_schema, walk_id_for
from keywalk_audit.walks.fingerprint import geometric_fingerprint
from keywalk_audit.walks.scorer import score_walk

FIXTURE_PWDUMP = Path(__file__).parent / "fixtures" / "sample_pwdump.txt"

WALK_PLAINTEXTS: tuple[str, ...] = (
    "1qazxsw2!QAZXSW@",
    "1qaz2wsx!QAZ@WSX",
    "zaq12wsxZAQ!@WSX",
    "zaq1xsw2ZAQ!XSW@",
    "1qaz2wsx3edc4rfv",
    "4rfv3edc2wsx1qaz",
    "1qazxcde3!QAZXCDE#",
    "3edcxzaq1#EDCXZAQ!",
    "qwertyui",
    "1qaz2wsx",
)


def _populate_rainbow(db_path: Path, plaintexts: tuple[str, ...]) -> None:
    """Insert the given plaintexts into a fresh rainbow at db_path."""
    with closing(duckdb.connect(str(db_path))) as conn:
        init_schema(conn)
        for text in plaintexts:
            wid = walk_id_for(QWERTY_US.name, text)
            score = score_walk(text, QWERTY_US)
            fp = geometric_fingerprint(text, QWERTY_US)
            conn.execute(
                "INSERT INTO candidates VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP);",
                [
                    wid,
                    text,
                    QWERTY_US.name,
                    len(text),
                    score.total,
                    fp,
                    False,
                    False,
                    False,
                ],
            )
            for algo in ("ntlm", "lm"):
                conn.execute(
                    "INSERT INTO hashes (walk_id, algorithm, hash_value) VALUES (?, ?, ?);",
                    [wid, algo, compute_hash(algo, text)],
                )

    geom = GeometricMinHashIndex(num_perm=128, threshold=0.6)
    string = StringMinHashIndex(num_perm=128, threshold=0.6)
    for text in plaintexts:
        wid = walk_id_for(QWERTY_US.name, text)
        geom.add(wid, text, QWERTY_US)
        string.add(wid, text)
    geom.save(geom_lsh_path(db_path))
    string.save(string_lsh_path(db_path))


def test_audit_e2e_pwdump_finds_exactly_ten_walks(tmp_path: Path) -> None:
    """Acceptance criterion #4."""
    db = tmp_path / "rainbow.duckdb"
    _populate_rainbow(db, WALK_PLAINTEXTS)
    report = audit(FIXTURE_PWDUMP, db, threshold=0.7)
    assert report.total_accounts == 20
    assert report.walk_accounts == 10
    finding_users = {f.username for f in report.findings}
    expected = {f"walker{i:02d}" for i in range(1, 11)}
    assert finding_users == expected
    for f in report.findings:
        assert f.matched_algorithm in {"ntlm", "lm"}
        assert f.walk_score >= 0.7
        assert f.layout == "qwerty_us"


def test_audit_threshold_excludes_low_scores(tmp_path: Path) -> None:
    db = tmp_path / "rainbow.duckdb"
    _populate_rainbow(db, WALK_PLAINTEXTS)
    report = audit(FIXTURE_PWDUMP, db, threshold=0.99)
    assert report.walk_accounts == 0


def test_audit_returns_zero_when_rainbow_empty(tmp_path: Path) -> None:
    db = tmp_path / "rainbow.duckdb"
    with closing(duckdb.connect(str(db))) as conn:
        init_schema(conn)
    report = audit(FIXTURE_PWDUMP, db, threshold=0.7)
    assert report.total_accounts == 20
    assert report.walk_accounts == 0


def test_audit_reports_error_when_sam_missing(tmp_path: Path) -> None:
    db = tmp_path / "rainbow.duckdb"
    with closing(duckdb.connect(str(db))) as conn:
        init_schema(conn)
    report = audit(tmp_path / "missing.pwdump", db)
    assert report.errors
    assert report.total_accounts == 0


def test_write_json_round_trip(tmp_path: Path) -> None:
    db = tmp_path / "rainbow.duckdb"
    _populate_rainbow(db, WALK_PLAINTEXTS)
    report = audit(FIXTURE_PWDUMP, db, threshold=0.7)
    out = tmp_path / "report.json"
    write_json(report, out)
    parsed = json.loads(out.read_text(encoding="utf-8"))
    assert parsed["walk_accounts"] == 10
    assert len(parsed["findings"]) == 10


def test_write_xlsx_has_expected_sheets(tmp_path: Path) -> None:
    db = tmp_path / "rainbow.duckdb"
    _populate_rainbow(db, WALK_PLAINTEXTS)
    report = audit(FIXTURE_PWDUMP, db, threshold=0.7)
    out = tmp_path / "report.xlsx"
    write_xlsx(report, out)
    wb = load_workbook(out)
    assert {"Summary", "Findings", "Fuzzy_Clusters", "Score_Distribution"}.issubset(
        set(wb.sheetnames)
    )
    findings_sheet = wb["Findings"]
    rows = list(findings_sheet.iter_rows(values_only=True))
    assert rows[0] == (
        "username",
        "rid",
        "matched_algorithm",
        "plaintext",
        "walk_score",
        "layout",
        "fingerprint",
        "geom_cluster_size",
        "string_cluster_size",
    )
    assert len(rows) == 11  # header + 10 findings
