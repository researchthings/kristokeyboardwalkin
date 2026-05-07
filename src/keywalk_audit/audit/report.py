"""Audit report writers: JSON and XLSX.

The XLSX layout has four sheets:
- ``Summary``: total accounts, walk count, walk percentage, runtime.
- ``Findings``: one row per matched account.
- ``Fuzzy_Clusters``: one row per fuzzy cluster with member walk IDs.
- ``Score_Distribution``: histogram of finding walk scores.
"""

from __future__ import annotations

import json
from collections import Counter
from dataclasses import asdict
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook

if TYPE_CHECKING:
    from pathlib import Path

    from keywalk_audit.audit.runner import AuditFinding, AuditReport

_HISTOGRAM_BINS: tuple[tuple[float, float], ...] = (
    (0.7, 0.75),
    (0.75, 0.8),
    (0.8, 0.85),
    (0.85, 0.9),
    (0.9, 0.95),
    (0.95, 1.0001),
)


def _report_to_dict(report: AuditReport) -> dict[str, Any]:
    body = asdict(report)
    body["findings"] = [asdict(f) for f in report.findings]
    return body


def write_json(report: AuditReport, path: Path) -> None:
    payload = _report_to_dict(report)
    with path.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, sort_keys=True)


def _walk_percentage(report: AuditReport) -> float:
    if report.total_accounts == 0:
        return 0.0
    return 100.0 * report.walk_accounts / report.total_accounts


def _histogram(findings: tuple[AuditFinding, ...]) -> list[tuple[str, int]]:
    counter: Counter[str] = Counter()
    for f in findings:
        for low, high in _HISTOGRAM_BINS:
            if low <= f.walk_score < high:
                counter[f"{low:.2f}-{high:.2f}"] += 1
                break
    rows: list[tuple[str, int]] = []
    for low, high in _HISTOGRAM_BINS:
        label = f"{low:.2f}-{high:.2f}"
        rows.append((label, counter.get(label, 0)))
    return rows


def write_xlsx(report: AuditReport, path: Path) -> None:
    wb = Workbook()
    summary = wb.active
    if summary is None:
        msg = "openpyxl returned no active sheet"
        raise RuntimeError(msg)
    summary.title = "Summary"
    summary.append(["Metric", "Value"])
    summary.append(["Total accounts", report.total_accounts])
    summary.append(["Walk accounts", report.walk_accounts])
    summary.append(["Walk percentage", f"{_walk_percentage(report):.2f}%"])
    summary.append(["Threshold", report.threshold])
    summary.append(["Runtime seconds", f"{report.runtime_seconds:.4f}"])
    if report.errors:
        summary.append(["Errors", "; ".join(report.errors)])

    findings_sheet = wb.create_sheet("Findings")
    findings_sheet.append(
        [
            "username",
            "rid",
            "matched_algorithm",
            "plaintext",
            "walk_score",
            "layout",
            "fingerprint",
            "geom_cluster_size",
            "string_cluster_size",
        ]
    )
    for f in report.findings:
        findings_sheet.append(
            [
                f.username,
                f.rid,
                f.matched_algorithm,
                f.plaintext,
                f.walk_score,
                f.layout,
                f.fingerprint,
                len(f.fuzzy_geom_cluster),
                len(f.fuzzy_str_cluster),
            ]
        )

    clusters_sheet = wb.create_sheet("Fuzzy_Clusters")
    clusters_sheet.append(["username", "kind", "member_walk_id"])
    for f in report.findings:
        for member in f.fuzzy_geom_cluster:
            clusters_sheet.append([f.username, "geometric", member])
        for member in f.fuzzy_str_cluster:
            clusters_sheet.append([f.username, "string", member])

    histogram_sheet = wb.create_sheet("Score_Distribution")
    histogram_sheet.append(["score_bin", "count"])
    for label, count in _histogram(report.findings):
        histogram_sheet.append([label, count])

    wb.save(str(path))
